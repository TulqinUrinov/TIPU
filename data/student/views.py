import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side
from rest_framework import generics, filters, status
from rest_framework.exceptions import PermissionDenied
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework import filters

from django.http import HttpResponse

from data.faculty.models import Faculty
from data.student.services import StudentFilterService
from sms.sayqal import SayqalSms
from data.common.pagination import CustomPagination
from data.common.permission import IsAuthenticatedUserType

from data.student.serializers import *

# O'quv yiliga tegishli barcha talabalar ro'yxati uchun
from django.db.models import Value, F, DecimalField, ExpressionWrapper, Avg, Sum, Case, IntegerField, When, Count, Q


# Student List
class StudentEduYearListApiView(generics.ListAPIView):
    serializer_class = StudentEduYearSerializer
    pagination_class = CustomPagination
    permission_classes = [IsAuthenticatedUserType]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = [
        "full_name",
        "jshshir",
        "user_account__phone_number",
        "phone_number"
    ]

    ordering_fields = ["full_name", "left"]  # faqat shu maydonlar bo‘yicha sortlashga ruxsat

    def get_queryset(self):
        return StudentFilterService.filter_students(self.request, self.kwargs.get('edu_year'))

    # def get_queryset(self):
    #     edu_year = self.kwargs.get('edu_year')
    #     course = self.request.query_params.get('course')
    #     faculty_ids = self.request.query_params.get('faculty')
    #     specialization_ids = self.request.query_params.get('specialization')
    #     percentage_range = self.request.query_params.get('percentage')
    #     type_filter = self.request.query_params.get('type')
    #     status = self.request.query_params.get('status')
    #     education_form = self.request.query_params.get('education_form')

    # queryset = Student.objects.filter(
    #     student_years__education_year_id=edu_year
    # )

    # queryset = Student.objects.filter(
    #     student_years__education_year_id=edu_year
    # ).annotate(
    #     contract_amount=Coalesce(
    #         F("contract__period_amount_dt"),
    #         Value(0, output_field=DecimalField(max_digits=15, decimal_places=2))
    #     ),
    #     left=Coalesce(
    #         Sum("contract_payments__left"),
    #         Value(0, output_field=DecimalField(max_digits=15, decimal_places=2))
    #     ),
    # ).annotate(
    #     total_paid=F("contract_amount") - F("left"),
    #     percentage=ExpressionWrapper(
    #         (F("total_paid") * Value(100, output_field=DecimalField()))
    #         / NullIf(F("contract_amount"), Value(0, output_field=DecimalField())),
    #         output_field=DecimalField(max_digits=5, decimal_places=2)
    #     )
    # )
    #
    # # Kurs bo‘yicha filter
    # if course:
    #     queryset = queryset.filter(course=course)
    #
    # # Student statusi bo'yicha filter
    # if status:
    #     queryset = queryset.filter(status=status)
    #
    # if education_form:
    #     queryset = queryset.filter(education_form=education_form)
    #
    # # Fakultet bo‘yicha filter
    # if faculty_ids:
    #     faculty_list = [int(f_id) for f_id in faculty_ids.split(",")]
    #     queryset = queryset.filter(specialization__faculty_id__in=faculty_list)
    #
    # # Yo'nalishi bo'yicha filter
    # if specialization_ids:
    #     specialization_list = [int(s_id) for s_id in specialization_ids.split(",")]
    #     queryset = queryset.filter(specialization_id__in=specialization_list)
    #
    # # HEMIS / NO-HEMIS bo‘yicha filter
    # if type_filter == "hemis":
    #     queryset = queryset.filter(user_account__isnull=True)
    # if type_filter == "no-hemis":
    #     queryset = queryset.filter(user_account__isnull=False)
    #
    # # percentage bo‘yicha filter
    # if percentage_range:
    #     if "-" in percentage_range:
    #         start, end = map(float, percentage_range.split("-"))
    #         queryset = queryset.filter(
    #             percentage__gte=start,
    #             percentage__lte=end
    #         )
    #     else:
    #         value = float(percentage_range)
    #         queryset = queryset.filter(percentage=value)
    #
    # return queryset.distinct()


# Student List Excel
class StudentEduYearExcelExportApiView(generics.GenericAPIView):
    permission_classes = [IsAuthenticatedUserType]

    def get_queryset(self):
        edu_year = self.kwargs.get('edu_year')
        course = self.request.query_params.get('course')
        faculty_ids = self.request.query_params.get('faculty')
        percentage_range = self.request.query_params.get('percentage')
        type_filter = self.request.query_params.get('type')
        print(type_filter)

        queryset = Student.objects.filter(
            student_years__education_year_id=edu_year
        ).annotate(
            contract_amount=Coalesce(
                F("contract__period_amount_dt"),
                Value(0, output_field=DecimalField(max_digits=15, decimal_places=2))
            ),
            left=Coalesce(
                Sum("contract_payments__left"),
                Value(0, output_field=DecimalField(max_digits=15, decimal_places=2))
            ),
        ).annotate(
            total_paid=F("contract_amount") - F("left"),
            percentage=ExpressionWrapper(
                (F("total_paid") * Value(100, output_field=DecimalField()))
                / NullIf(F("contract_amount"), Value(0, output_field=DecimalField())),
                output_field=DecimalField(max_digits=5, decimal_places=2)
            )
        )

        if course:
            queryset = queryset.filter(course=course)

        if faculty_ids:
            faculty_list = [int(f_id) for f_id in faculty_ids.split(",")]
            queryset = queryset.filter(specialization__faculty_id__in=faculty_list)

        if type_filter == "hemis":
            queryset = queryset.filter(user_account__isnull=True)
        if type_filter == "no-hemis":
            print(type_filter)
            queryset = queryset.filter(user_account__isnull=False)

        if percentage_range:
            start, end = map(float, percentage_range.split("-"))
            queryset = queryset.filter(
                percentage__gte=start,
                percentage__lte=end
            )

        return queryset.distinct()

    def get(self, request, *args, **kwargs):
        queryset = self.get_queryset()

        # Excel fayl yaratamiz
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Talabalar"

        # Header qator
        headers = [
            "F.I.Sh", "JSHSHIR", "Telefon",
            "Fakultet", "Mutaxassislik", "Guruh",
            "Kontrakt summasi", "To‘langan", "Qoldiq", "Foiz (%)"
        ]
        ws.append(headers)

        # Ma'lumotlarni yozish
        for student in queryset:
            print(student)
            ws.append([
                student.full_name,
                student.jshshir,
                student.user_account.phone_number if hasattr(student,
                                                             "user_account") and student.user_account else student.phone_number,
                student.specialization.faculty.name if student.specialization and student.specialization.faculty else "",
                student.specialization.name if student.specialization else "",
                student.group,
                float(student.contract_amount or 0),
                float(student.total_paid or 0),
                float(student.left or 0),
                float(student.percentage or 0),
            ])

        # Javobga yozib yuboramiz
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="students.xlsx"'
        wb.save(response)
        return response


# Retrieve
class StudentDetailApiView(generics.RetrieveUpdateAPIView):
    queryset = Student.objects.filter(is_archived=False)
    serializer_class = StudentSerializer
    permission_classes = [IsAuthenticatedUserType]

    def get_object(self):
        if getattr(self.request, "admin_user", None):
            return super().get_object()

        if getattr(self.request, "student_user", None):
            student = Student.objects.filter(
                user_account=self.request.student_user,
                is_archived=False
            ).first()
            if not student:
                raise PermissionDenied("Siz faqat o‘zingizni ko‘rishingiz mumkin")
            return student

        raise PermissionDenied("Ruxsat yo‘q")


# Statistics
class StudentStatisticsApiView(APIView):
    permission_classes = [IsAuthenticatedUserType]

    def get(self, request, edu_year):
        course = request.query_params.get("course")  # masalan: ?course=1-kurs
        faculty_ids = request.query_params.get("faculty")  # masalan: ?faculty=1,2,3
        specialization_ids = self.request.query_params.get('specialization')
        student_status = request.query_params.get("status")
        education_form = self.request.query_params.get('education_form')

        filters = {"is_archived": False}

        if course:
            filters["course"] = course

        if student_status:
            filters["status"] = student_status

        if education_form:
            filters["education_form"] = education_form

        if faculty_ids:
            faculty_list = [int(f_id) for f_id in faculty_ids.split(",") if f_id.isdigit()]
            filters["specialization__faculty_id__in"] = faculty_list

        if specialization_ids:
            specialization_list = [int(s_id) for s_id in specialization_ids.split(",")]
            filters["specialization_id__in"] = specialization_list

        serializer = StudentStatisticsSerializer(
            instance=Student(),
            context={
                "filters": filters,
                "request": request,
                "edu_year": edu_year,
            }
        )
        return Response(serializer.data, status=status.HTTP_200_OK)


# Statistics Excel
class StatisticsExcelApiView(APIView):
    permission_classes = [IsAuthenticatedUserType]

    def get(self, request, edu_year):
        course = request.query_params.get("course")
        faculty_ids = request.query_params.get("faculty")
        percentage_range = request.query_params.get("percentage")

        filters = {"is_archived": False, "student_years__education_year_id": edu_year}
        if course:
            filters["course"] = course
        if faculty_ids:
            faculty_list = [int(f_id) for f_id in faculty_ids.split(",") if f_id.isdigit()]
            faculties = Faculty.objects.filter(id__in=faculty_list)
        else:
            faculties = Faculty.objects.all()

        students = Student.objects.filter(**filters).annotate(
            contract_amount=Coalesce(
                F("contract__period_amount_dt"),
                Value(0, output_field=DecimalField(max_digits=15, decimal_places=2))
            ),
            left=Coalesce(
                Sum("contract_payments__left"),
                Value(0, output_field=DecimalField(max_digits=15, decimal_places=2))
            ),
        ).annotate(
            total_paid=F("contract_amount") - F("left"),
            percentage=(F("contract_amount") - F("left")) * 100 /
                       NullIf(F("contract_amount"),
                              Value(0, output_field=DecimalField(max_digits=15, decimal_places=2)))
        )

        if percentage_range:
            if "-" in percentage_range:
                start, end = map(float, percentage_range.split("-"))
                percentage_filter = {"percentage__gte": start, "percentage__lte": end}
                percentage_label = f"{start}–{end} foiz oralig‘ida to‘laganlar (soni)"
            else:
                value = float(percentage_range)
                percentage_filter = {"percentage": value}
                percentage_label = f"{value} foiz to‘laganlar (soni)"
        else:
            # Default — faqat haqiqatan ham to‘lov qilganlar
            percentage_filter = {"percentage__gt": 0, "percentage__lte": 100}
            percentage_label = "0–100 foiz oralig‘ida to‘laganlar (soni)"

        jami_shartnomalar = 0
        jami_tolaganlar = 0
        jami_foiz = 0
        jami_foiz_soni = 0

        # Excel yaratish
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Statistika"

        # Dinamik sarlavha
        ws.append(
            ["№", "Dekanatlar", "Tuzilgan shartnomalar soni", percentage_label, "O‘rtacha foiz"]
        )
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Fakultetlar bo‘yicha
        for i, faculty in enumerate(faculties, start=1):
            faculty_students = students.filter(specialization__faculty=faculty)
            total_contracts = faculty_students.count()

            if percentage_filter:
                filtered_students = faculty_students.filter(**percentage_filter)
            else:
                filtered_students = faculty_students.filter(percentage__gte=100)

            total_paid_students = filtered_students.count()

            # o‘rtacha foiz (agar talabalar bo‘lsa)
            avg_percentage = filtered_students.aggregate(avg=Avg("percentage"))["avg"] or 0

            jami_shartnomalar += total_contracts
            jami_tolaganlar += total_paid_students
            jami_foiz += avg_percentage * total_paid_students
            jami_foiz_soni += total_paid_students

            ws.append([
                i,
                faculty.name,
                total_contracts,
                total_paid_students,
                f"{round(avg_percentage, 2)}%" if total_paid_students > 0 else "0%"
            ])

        # Jami qatori (umumiy o‘rtacha foiz)
        umumiy_avg = jami_foiz / jami_foiz_soni if jami_foiz_soni > 0 else 0

        ws.append([
            "",
            "Jami",
            jami_shartnomalar,
            jami_tolaganlar,
            f"{round(umumiy_avg, 2)}%"
        ])

        # Response qaytarish
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = 'attachment; filename=faculty_statistics.xlsx'
        wb.save(response)
        return response


# Fakultet bo'yicha studentlar statistikasi
class FacultyStatsAPIView(APIView):
    permission_classes = [IsAuthenticatedUserType]
    """
    Fakultetlar kesimida talabalar kontrakt to'lovi statistikasi
    """

    def get(self, request, edu_year):
        # Avval subquery yordamida har bir talaba uchun to'lov ma'lumotlarini olamiz
        from django.db.models import Subquery, OuterRef

        # Subquery: har bir talaba uchun qolgan to'lov miqdorini hisoblaymiz
        total_left_subquery = Student.objects.filter(
            id=OuterRef('id')
        ).annotate(
            total_left=Coalesce(
                Sum("contract_payments__left"),
                Value(0, output_field=DecimalField(max_digits=15, decimal_places=2))
            )
        ).values('total_left')[:1]

        # Asosiy queryset
        students = Student.objects.filter(
            student_years__education_year_id=edu_year
        ).annotate(
            contract_amount=Coalesce(
                F("contract__period_amount_dt"),
                Value(0, output_field=DecimalField(max_digits=15, decimal_places=2))
            ),
            total_left=Subquery(total_left_subquery),
            total_paid=ExpressionWrapper(
                F("contract_amount") - F("total_left"),
                output_field=DecimalField(max_digits=15, decimal_places=2)
            ),
            percentage=Case(
                When(contract_amount=0, then=Value(0)),
                default=ExpressionWrapper(
                    (F("total_paid") * 100) / F("contract_amount"),
                    output_field=DecimalField(max_digits=5, decimal_places=2)
                ),
                output_field=DecimalField(max_digits=5, decimal_places=2)
            )
        )

        # Endi fakultetlar bo'yicha statistikani hisoblaymiz
        stats = students.values(
            "specialization__faculty__name"
        ).annotate(
            total=Count("id"),
            fully_paid=Count("id", filter=Q(percentage__gte=100)),
            p76_99=Count("id", filter=Q(percentage__gte=76, percentage__lt=100)),
            p50_75=Count("id", filter=Q(percentage__gte=50, percentage__lt=76)),
            p25_49=Count("id", filter=Q(percentage__gte=25, percentage__lt=50)),
            p0_24=Count("id", filter=Q(percentage__lt=25)),
        )

        total_students = Student.objects.filter(is_archived=False).count()

        # Foizlarni ham qo'shib chiqamiz
        results = []
        results.append({
            "total_students": total_students,
        })
        for item in stats:
            total = item["total"]
            faculty = item["specialization__faculty__name"]
            results.append({
                "faculty": faculty,
                "total": total,
                "fully_paid": {
                    "count": item["fully_paid"],
                    "percent": round((item["fully_paid"] / total) * 100, 1) if total else 0
                },
                "p76_99": {
                    "count": item["p76_99"],
                    "percent": round((item["p76_99"] / total) * 100, 1) if total else 0
                },
                "p50_75": {
                    "count": item["p50_75"],
                    "percent": round((item["p50_75"] / total) * 100, 1) if total else 0
                },
                "p25_49": {
                    "count": item["p25_49"],
                    "percent": round((item["p25_49"] / total) * 100, 1) if total else 0
                },
                "p0_24": {
                    "count": item["p0_24"],
                    "percent": round((item["p0_24"] / total) * 100, 1) if total else 0
                },
            })

        return Response(results, status=status.HTTP_200_OK)


# Send SMS
class SendSmsView(APIView):
    permission_classes = [IsAuthenticatedUserType]

    def post(self, request, edu_year=None):
        serializer = SendSmsSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        message = serializer.validated_data['message']
        jshshir_list = serializer.validated_data.get("students", [])
        send_all = serializer.validated_data.get("send_all", False)
        send_filtered = serializer.validated_data.get("send_filtered", False)

        sms_client = SayqalSms()
        success, failed = [], []

        if send_all:
            students = Student.objects.all()
        elif send_filtered:
            students = StudentFilterService.filter_students(request, edu_year)
        else:
            students = Student.objects.filter(jshshir__in=jshshir_list)

        for student in students:
            student_user = getattr(student, "user_account", None)
            phone_number = (
                student_user.phone_number if student_user and student_user.phone_number
                else student.phone_number
            )
            if phone_number:
                response = sms_client.send_sms(phone_number, message)
                if response.status_code == status.HTTP_200_OK:
                    success.append({"jshshir": student.jshshir, "student": student.full_name})
                else:
                    failed.append({
                        "jshshir": student.jshshir,
                        "student": student.full_name,
                        "error": response.text
                    })
            else:
                failed.append({
                    "jshshir": student.jshshir,
                    "student": student.full_name,
                    "error": "Phone number not found."
                })

        return Response({"success": success, "failed": failed}, status=status.HTTP_200_OK)

# # Send Sms to choosen students
# class SendSmsView(APIView):
#     permission_classes = [IsAuthenticatedUserType]
#
#     def post(self, request):
#         serializer = SendSmsSerializer(data=request.data)
#         serializer.is_valid(raise_exception=True)
#
#         message = serializer.validated_data['message']
#         jshshir_list = serializer.validated_data.get("students", [])
#         send_all = serializer.validated_data.get("send_all", False)  # ✅ qo‘shimcha flag
#
#         sms_client = SayqalSms()
#         success, failed = [], []
#
#         # Agar frontend "send_all": true yuborsa → barcha studentlarni olish
#         if send_all:
#             students = Student.objects.all()
#         else:
#             students = Student.objects.filter(jshshir__in=jshshir_list)
#
#         for student in students:
#
#             student_user = getattr(student, "user_account", None)
#             phone_number = None
#
#             if student_user and student_user.phone_number:
#                 phone_number = student_user.phone_number
#             elif student.phone_number:
#                 phone_number = student.phone_number
#
#             if phone_number:
#                 response = sms_client.send_sms(phone_number, message)
#                 if response.status_code == status.HTTP_200_OK:
#                     success.append({"jshshir": student.jshshir, "student": student.full_name})
#                 else:
#                     failed.append({
#                         "jshshir": student.jshshir,
#                         "student": student.full_name,
#                         "error": response.text
#                     })
#             else:
#                 failed.append({
#                     "jshshir": student.jshshir,
#                     "student": student.full_name,
#                     "error": "Phone number not found."
#                 })
#
#         return Response({
#             "success": success,
#             "failed": failed
#         }, status=status.HTTP_200_OK)
